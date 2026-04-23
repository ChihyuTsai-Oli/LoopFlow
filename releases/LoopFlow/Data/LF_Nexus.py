# -*- coding: utf-8 -*-
# Script: LF_Nexus.py (LoopFlow 3D Core Engine)
# Version: v1.0
# Date: 2026-04-20
# Developer: Cursor (Claude Opus 4.7 / Sonnet 4.6)
# Environment: Rhino 8 (CPython 3.9) / Windows 10
# Synced Files: _LoopFlow_Config.py, _LF_Debug.py, LF_Push_3D_to_JSON.py, _LF_Registry.py
# Usage: Provides four main functions (selected via Rhino dialog):
#           1. Dict. to Layer  — Reads LoopFlow_Dictionary.xlsx and builds the M3D layer structure
#           2. Tag Trigger     — Scans M3D objects and writes elevation/space/UUID and all UserText attributes
#           3. Tag Checker     — Validates UUID format and uniqueness for all M3D objects
#           4. Rhino to XLSX   — Reverse-exports existing object attributes and updates the dictionary
#           Run in 3D.3dm. After Tag Trigger you may immediately push Project_Registry.json.

# ==================================================================
# Imports
# ==================================================================
import System
import sys
import scriptcontext as sc
import rhinoscriptsyntax as rs
import Rhino
import Eto.Forms as forms
import Eto.Drawing as drawing
import pandas as pd
import openpyxl
import os
import uuid
import importlib

# ==================================================================
# Config Reload & Aliases
# ==================================================================
import _LoopFlow_Config as _CFG
importlib.reload(_CFG)

DICTIONARY_FILENAME_XLSX = _CFG.DICTIONARY_FILENAME_XLSX
DICTIONARY_KEY_COLUMN    = _CFG.DICTIONARY_KEY_COLUMN
DICTIONARY_SKIPROWS      = _CFG.DICTIONARY_SKIPROWS
LAYER_PREFIX_3D          = _CFG.LAYER_PREFIX_3D
LAYER_DATA_SUFFIX        = _CFG.LAYER_DATA_SUFFIX
LAYER_SPACE_BOUNDARIES   = _CFG.LAYER_SPACE_BOUNDARIES
LAYER_LEVEL_FFL          = _CFG.LAYER_LEVEL_FFL
LAYER_LEVEL_FL           = _CFG.LAYER_LEVEL_FL
LAYER_DW_PLAN            = _CFG.LAYER_DW_PLAN
LAYER_CABINET_PREFIX     = _CFG.LAYER_CABINET_PREFIX
WHITE_LIST               = _CFG.WHITE_LIST
COLOR_LAYER_MAP          = _CFG.COLOR_LAYER_MAP
COLOR_DATA_LAYER         = getattr(_CFG, "COLOR_DATA_LAYER", (0, 0, 0))

try:
    from _LF_Debug import log_exception, DEBUG_LOG_PATH as _DEBUG_LOG_PATH
except Exception:
    log_exception = None
    _DEBUG_LOG_PATH = "cursor_LF_debug_log.txt"

# ==================================================================
# Constants
# ==================================================================
VALID_GEOM_TYPES = [8, 16, 32, 4096, 1073741824]
_PREFIX_WITH_SEP = LAYER_PREFIX_3D + "::"
_DATA_PATH_FRAGMENT = "::" + LAYER_DATA_SUFFIX + "::"
_DATA_PATH_PREFIX   = LAYER_DATA_SUFFIX + "::"

# ==================================================================
# Layer & Color Utilities
# ==================================================================
def _to_color(c):
    if isinstance(c, (list, tuple)):
        return System.Drawing.Color.FromArgb(int(c[0]), int(c[1]), int(c[2]))
    h = str(c).lstrip('#')
    return System.Drawing.Color.FromArgb(int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))

def _build_full_layer_path(rel_path):
    if rel_path.startswith(_PREFIX_WITH_SEP):
        return rel_path
    return _PREFIX_WITH_SEP + rel_path

def _strip_prefix(full_path):
    if full_path.startswith(_PREFIX_WITH_SEP):
        return full_path[len(_PREFIX_WITH_SEP):]
    return full_path

def _is_data_layer(layer_name):
    if not layer_name: return False
    if layer_name == LAYER_PREFIX_3D + "::" + LAYER_DATA_SUFFIX: return True
    return _DATA_PATH_FRAGMENT in layer_name or layer_name.startswith(_DATA_PATH_PREFIX)

# ==================================================================
# Geometry & Data Helpers
# ==================================================================
def format_elevation_val(val):
    formatted = "{:+,.1f}".format(val)
    if formatted in ["+0.0", "-0.0"]: return u"\u00b10"
    return formatted

def get_dictionary_path(force_select=False):
    doc = Rhino.RhinoDoc.ActiveDoc
    if doc and doc.Path:
        proj_dir = os.path.dirname(doc.Path)
        local_xlsx = os.path.join(proj_dir, DICTIONARY_FILENAME_XLSX)
        if not force_select and os.path.exists(local_xlsx):
            return local_xlsx

    doc_path = rs.DocumentPath()
    file_path = rs.OpenFileName(u"Select LoopFlow Dictionary", u"LoopFlow Dictionary (*.xlsx)|*.xlsx||", folder=doc_path)
    return file_path

def load_dict_from_path(file_path):
    if not file_path: return None
    try:
        df = pd.read_excel(file_path, skiprows=DICTIONARY_SKIPROWS)
        df.columns = [str(c).strip() for c in df.columns]

        if DICTIONARY_KEY_COLUMN not in df.columns:
            rs.MessageBox(u"Column not found in dictionary: '{}'\nPlease confirm Excel structure: row 1 = version header, row 2 = column headers.".format(DICTIONARY_KEY_COLUMN), 48)
            return None

        df = df.dropna(subset=[DICTIONARY_KEY_COLUMN])
        df[DICTIONARY_KEY_COLUMN] = df[DICTIONARY_KEY_COLUMN].astype(str).str.strip()
        cols = [c for c in df.columns if str(c).startswith("_") or c == DICTIONARY_KEY_COLUMN]
        return df[cols].fillna("-").replace(["", "nan", "None"], "-")

    except Exception as e:
        rs.MessageBox(u"Unexpected error while parsing dictionary:\n{}".format(e), 16)
        return None

def get_closest_curve_dist_2d(pt_3d, crv_id):
    z_crv = rs.CurveStartPoint(crv_id).Z
    pt_test = Rhino.Geometry.Point3d(pt_3d.X, pt_3d.Y, z_crv)
    param = rs.CurveClosestPoint(crv_id, pt_test)
    if param is not None:
        pt_on = rs.EvaluateCurve(crv_id, param)
        return pt_test.DistanceTo(pt_on)
    return float('inf')

def get_dimensions(obj_id):
    rhino_obj = rs.coercerhinoobject(obj_id)
    if not rhino_obj: return "-", "-", "-"

    if isinstance(rhino_obj, Rhino.DocObjects.InstanceObject):
        xform = rhino_obj.InstanceXform
        idef  = rhino_obj.InstanceDefinition
        geom_list = idef.GetObjects()
        if not geom_list: return "-", "-", "-"

        bbox = Rhino.Geometry.BoundingBox.Empty
        for geom in geom_list:
            g_bbox = geom.Geometry.GetBoundingBox(True)
            if g_bbox.IsValid: bbox.Union(g_bbox)

        if not bbox.IsValid: return "-", "-", "-"

        base_w = bbox.Max.X - bbox.Min.X
        base_d = bbox.Max.Y - bbox.Min.Y
        base_h = bbox.Max.Z - bbox.Min.Z

        vec_x = Rhino.Geometry.Vector3d(1, 0, 0)
        vec_y = Rhino.Geometry.Vector3d(0, 1, 0)
        vec_z = Rhino.Geometry.Vector3d(0, 0, 1)
        vec_x.Transform(xform)
        vec_y.Transform(xform)
        vec_z.Transform(xform)

        w = base_w * vec_x.Length
        d = base_d * vec_y.Length
        h = base_h * vec_z.Length
    else:
        pts = rs.BoundingBox(obj_id)
        if not pts: return "-", "-", "-"
        w = pts[0].DistanceTo(pts[1])
        d = pts[0].DistanceTo(pts[3])
        h = pts[0].DistanceTo(pts[4])

    return "{:.1f}".format(w), "{:.1f}".format(d), "{:.1f}".format(h)

def get_elevation_value(obj_id, basis):
    basis = str(basis).upper()
    if basis not in ["BH", "TH", "BC", "TH/BH", "CH"]: return "-"

    bbox = rs.BoundingBox(obj_id)
    if not bbox: return "-"
    obj_th, obj_bh = bbox[4].Z, bbox[0].Z
    obj_mid = (obj_th + obj_bh) / 2.0

    slab_centers = []
    for l in rs.LayerNames():
        if "00_STR" in l.upper() or "SLAB" in l.upper() or u"\u6a13\u677f" in l:
            oids = rs.ObjectsByLayer(l)
            if oids:
                for oid in oids:
                    s_bbox = rs.BoundingBox(oid)
                    if s_bbox: slab_centers.append((s_bbox[4].Z + s_bbox[0].Z) / 2.0)
    slab_centers = sorted(list(set(slab_centers)))

    boundary_layers = [LAYER_LEVEL_FFL, LAYER_LEVEL_FL]
    all_b_ids = []
    for l in boundary_layers:
        if rs.IsLayer(l): all_b_ids.extend(rs.ObjectsByLayer(l))

    cp = Rhino.Geometry.Point3d((bbox[0].X+bbox[2].X)/2, (bbox[0].Y+bbox[2].Y)/2, obj_mid)
    all_oframes = []
    for b_id in all_b_ids:
        if rs.IsCurve(b_id) and rs.IsCurveClosed(b_id):
            curve_z  = rs.CurveStartPoint(b_id).Z
            cp_2d    = Rhino.Geometry.Point3d(cp.X, cp.Y, curve_z)
            if rs.coercecurve(b_id).Contains(cp_2d, Rhino.Geometry.Plane.WorldXY, 0.01) == Rhino.Geometry.PointContainment.Inside:
                dist = 0.0
            else:
                dist = get_closest_curve_dist_2d(cp_2d, b_id)
            if dist < 2000.0:
                try:
                    datum = float(rs.ObjectName(b_id) or rs.GetUserText(b_id, "Space_Name"))
                    all_oframes.append({"id": b_id, "z": curve_z, "datum": datum, "dist": dist})
                except: continue

    if not all_oframes: return "-"
    all_oframes.sort(key=lambda x: x['dist'])
    close_oframes = [o for o in all_oframes if o['dist'] <= all_oframes[0]['dist'] + 50.0]

    final_candidates = []
    for o in close_oframes:
        s_bottom, s_top = -999999.0, 999999.0
        for sz in slab_centers:
            if sz <= o['z'] + 200.0: s_bottom = sz
            if sz > o['z'] + 200.0: s_top = sz; break
        if s_bottom - 1.0 <= obj_mid <= s_top + 1.0:
            if o['z'] <= obj_th + 1.0: final_candidates.append(o)

    if not final_candidates:
        final_candidates = [o for o in close_oframes if o['z'] <= obj_th + 1.0]

    if final_candidates:
        final_candidates.sort(key=lambda x: x['z'], reverse=True)
        o_frame_datum = final_candidates[0]['datum']
    else:
        o_frame_datum = close_oframes[0]['datum']

    if basis == "TH/BH":
        rel_th, rel_bh = obj_th - o_frame_datum, obj_bh - o_frame_datum
        if abs(rel_th) <= abs(rel_bh): return u"{} / -".format(format_elevation_val(rel_th))
        else: return u"- / {}".format(format_elevation_val(rel_bh))

    calc_z = obj_bh
    if basis == "TH": calc_z = obj_th
    elif basis == "BC" and rs.IsBlockInstance(obj_id): calc_z = rs.BlockInstanceInsertPoint(obj_id).Z
    return format_elevation_val(calc_z - o_frame_datum)

def get_space_name_at_object(obj_id):
    if not rs.IsLayer(LAYER_SPACE_BOUNDARIES): return "EXT"
    bbox = rs.BoundingBox(obj_id)
    if not bbox: return "EXT"
    cp = Rhino.Geometry.Point3d((bbox[0].X+bbox[2].X)/2, (bbox[0].Y+bbox[2].Y)/2, bbox[0].Z)
    boundary_ids = rs.ObjectsByLayer(LAYER_SPACE_BOUNDARIES)
    if not boundary_ids: return "EXT"
    for b_id in boundary_ids:
        if rs.IsCurve(b_id) and rs.IsCurveClosed(b_id):
            if rs.coercecurve(b_id).Contains(cp, Rhino.Geometry.Plane.WorldXY, 0.01) == Rhino.Geometry.PointContainment.Inside:
                return rs.GetUserText(b_id, "Space_Name") or rs.ObjectName(b_id) or "EXT"
    return "EXT"

# ==================================================================
# JSON Push Bridge
# ==================================================================
def execute_push_to_json():
    try:
        script_dir = os.path.dirname(os.path.realpath(__file__))
        if script_dir not in sys.path: sys.path.append(script_dir)

        import _LF_Registry as _REG
        importlib.reload(_REG)

        import LF_Push_3D_to_JSON as _PUSH
        importlib.reload(_PUSH)
        _PUSH.push_3d_data()
    except Exception as e:
        if log_exception:
            print(log_exception(u"LF_Nexus.execute_push_to_json", e))
        rs.MessageBox(u"Failed to call registry push.\nError: {}\n\nSee debug log: {}".format(e, _DEBUG_LOG_PATH), 16)

# ==================================================================
# UUID Validation
# ==================================================================
_UUID_PATTERN = None
def _is_valid_uuid(s):
    global _UUID_PATTERN
    if _UUID_PATTERN is None:
        import re as _re
        _UUID_PATTERN = _re.compile(r"^[0-9A-Fa-f]{8}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{12}$")
    if not s: return False
    return bool(_UUID_PATTERN.match(str(s).strip()))

def check_global_uuids():
    all_objs = rs.AllObjects(True, True) or []
    used_uuids = {}
    duplicate_objs = set()
    invalid_objs   = set()
    for o in all_objs:
        u_val = rs.GetUserText(o, "_12_UUID")
        if not u_val or not u_val.strip():
            continue
        u_clean = u_val.strip()
        if not _is_valid_uuid(u_clean):
            invalid_objs.add(o)
            continue
        if u_clean in used_uuids:
            duplicate_objs.add(o)
            duplicate_objs.add(used_uuids[u_clean])
        else:
            used_uuids[u_clean] = o
    return duplicate_objs, invalid_objs

def _check_level_boundaries():
    valid_count   = 0
    invalid_names = []
    for l in [LAYER_LEVEL_FFL, LAYER_LEVEL_FL]:
        if not rs.IsLayer(l): continue
        ids = rs.ObjectsByLayer(l) or []
        for oid in ids:
            if not (rs.IsCurve(oid) and rs.IsCurveClosed(oid)):
                continue
            raw_name = rs.ObjectName(oid) or rs.GetUserText(oid, "Space_Name") or ""
            try:
                float(raw_name)
                valid_count += 1
            except (ValueError, TypeError):
                invalid_names.append((l, raw_name.strip() or u"(no name)"))
    return valid_count, invalid_names

# ==================================================================
# Tag Trigger: Write Attributes to Objects
# ==================================================================
def func_tag_trigger(df_dict):
    obj_ids = rs.AllObjects(True, True)
    if not obj_ids:
        rs.MessageBox(u"No objects found in the scene.", 48)
        return

    valid_count, invalid_names = _check_level_boundaries()
    if valid_count == 0:
        warn_msg  = u"Warning: No valid floor boundary curves found.\n\n"
        warn_msg += u"Ensure at least one closed curve exists on the following layers with its 'Object Name' set to an elevation number (e.g. 0, 305, 2600):\n"
        warn_msg += u"  · {}\n".format(LAYER_LEVEL_FFL)
        warn_msg += u"  · {}\n\n".format(LAYER_LEVEL_FL)
        if invalid_names:
            warn_msg += u"The following curves have non-numeric names and will be ignored:\n"
            for lname, cname in invalid_names[:10]:
                warn_msg += u"  \u00b7 [{}] Name='{}'\n".format(lname.split("::")[-1], cname)
            if len(invalid_names) > 10:
                warn_msg += u"  \u00b7 ... ({} more)\n".format(len(invalid_names) - 10)
            warn_msg += u"\n"
        warn_msg += u"If you continue, all _11_ Elevation fields will be written as '-'.\n\nContinue with TagTrigger?"
        if rs.MessageBox(warn_msg, 4 | 48, u"Floor Boundary Check") != 6:
            return

    layer_mapping = {str(row[DICTIONARY_KEY_COLUMN]).strip(): row for row in df_dict.to_dict(orient='records')}
    success_count  = 0
    new_uuid_count = 0

    duplicate_objs, invalid_uuid_objs = check_global_uuids()
    rs.EnableRedraw(False)

    for guid in obj_ids:
        layer_name = rs.ObjectLayer(guid)
        if not layer_name: continue
        if not layer_name.startswith(_PREFIX_WITH_SEP): continue
        if _is_data_layer(layer_name): continue
        if rs.ObjectType(guid) not in VALID_GEOM_TYPES: continue

        rel_layer = _strip_prefix(layer_name)

        if rel_layer in layer_mapping:
            tag_data = layer_mapping[rel_layer]

            current_uuid = rs.GetUserText(guid, "_12_UUID")
            needs_new_uuid = (
                not current_uuid or not current_uuid.strip()
                or guid in duplicate_objs
                or guid in invalid_uuid_objs
            )
            if needs_new_uuid:
                rs.SetUserText(guid, "_12_UUID", str(uuid.uuid4()).upper())
                new_uuid_count += 1

            basis_col = "_10_" if any(c.startswith("_10_") for c in df_dict.columns) else None
            if basis_col:
                actual_col = next((c for c in df_dict.columns if c.startswith("_10_")), None)
                basis = str(tag_data.get(actual_col, "BH")).upper() if actual_col else "BH"
            else:
                basis = "BH"

            elevation_val = get_elevation_value(guid, basis)
            space_val     = get_space_name_at_object(guid)
            dim_w, dim_d, dim_h = get_dimensions(guid)
            is_cb_layer   = LAYER_CABINET_PREFIX in layer_name

            for col in df_dict.columns:
                if col == DICTIONARY_KEY_COLUMN: continue

                if col.startswith("_01_"):
                    rs.SetUserText(guid, col, space_val)
                elif col.startswith("_12_"):
                    continue
                elif col.startswith("_11_"):
                    rs.SetUserText(guid, col, elevation_val)
                elif col.startswith("_05_"):
                    rs.SetUserText(guid, col, dim_w)
                elif col.startswith("_06_"):
                    rs.SetUserText(guid, col, dim_d)
                elif col.startswith("_07_"):
                    rs.SetUserText(guid, col, dim_h)
                elif col.startswith("_CB."):
                    if is_cb_layer:
                        existing = rs.GetUserText(guid, col)
                        if existing in [None, "", "-"]: rs.SetUserText(guid, col, str(tag_data.get(col, "-")))
                    else:
                        rs.SetUserText(guid, col, "-")
                else:
                    is_protected = any(col.startswith(wl[:4]) for wl in WHITE_LIST)
                    if is_protected:
                        if rs.GetUserText(guid, col) in [None, "", "-"]:
                            rs.SetUserText(guid, col, str(tag_data.get(col, "-")))
                    else:
                        rs.SetUserText(guid, col, str(tag_data.get(col, "-")))
            success_count += 1

    rs.UnselectAllObjects()
    rs.EnableRedraw(True)
    msg = u"Global TagTrigger write complete!\nProcessed {} objects.\n(Newly generated/repaired: {} UUIDs)\n\nPush to project registry (JSON) now?".format(success_count, new_uuid_count)
    if rs.MessageBox(msg, 4 | 64, u"Write Successful") == 6:
        execute_push_to_json()

# ==================================================================
# Tag Checker: Validate All Object Attributes
# ==================================================================
def func_tag_checker(df_dict):
    rs.UnselectAllObjects()
    objs = rs.AllObjects(True, True)
    if not objs:
        rs.UnselectAllObjects()
        return

    valid_count, invalid_names = _check_level_boundaries()
    if valid_count == 0:
        warn_msg  = u"Warning: No valid floor boundary curves found.\n\n"
        warn_msg += u"Ensure at least one closed curve exists on the following layers with its 'Object Name' set to an elevation number (e.g. 0, 305, 2600):\n"
        warn_msg += u"  · {}\n".format(LAYER_LEVEL_FFL)
        warn_msg += u"  · {}\n\n".format(LAYER_LEVEL_FL)
        if invalid_names:
            warn_msg += u"The following curves have non-numeric names and will be ignored:\n"
            for lname, cname in invalid_names[:10]:
                warn_msg += u"  \u00b7 [{}] Name='{}'\n".format(lname.split("::")[-1], cname)
            if len(invalid_names) > 10:
                warn_msg += u"  \u00b7 ... ({} more)\n".format(len(invalid_names) - 10)
            warn_msg += u"\n"
        warn_msg += u"If you continue, _11_ Elevation values cannot be accurately validated.\n\nContinue with TagChecker?"
        if rs.MessageBox(warn_msg, 4 | 48, u"Floor Boundary Check") != 6:
            return

    layer_mapping = {str(row[DICTIONARY_KEY_COLUMN]).strip(): row for row in df_dict.to_dict(orient='records')}
    red_list = []
    rs.EnableRedraw(False)

    duplicate_objs, invalid_uuid_objs = check_global_uuids()

    for guid in objs:
        layer_name = rs.ObjectLayer(guid)
        if not layer_name: continue
        if not layer_name.startswith(_PREFIX_WITH_SEP): continue
        if _is_data_layer(layer_name): continue
        if rs.ObjectType(guid) not in VALID_GEOM_TYPES: continue

        rel_layer = _strip_prefix(layer_name)
        if rel_layer not in layer_mapping: continue

        tag_data  = layer_mapping[rel_layer]
        is_error  = False

        current_uuid = rs.GetUserText(guid, "_12_UUID")
        if not current_uuid or not current_uuid.strip():
            is_error = True
        elif guid in invalid_uuid_objs:
            is_error = True
        elif guid in duplicate_objs:
            is_error = True

        if not is_error:
            is_cb_layer = LAYER_CABINET_PREFIX in rel_layer
            actual_col  = next((c for c in df_dict.columns if c.startswith("_10_")), None)
            basis       = str(tag_data.get(actual_col, "BH")).upper() if actual_col else "BH"

            true_elev   = get_elevation_value(guid, basis)
            true_space  = get_space_name_at_object(guid)
            dim_w, dim_d, dim_h = get_dimensions(guid)

            for col in df_dict.columns:
                if col == DICTIONARY_KEY_COLUMN: continue
                raw_val = rs.GetUserText(guid, col)
                if raw_val is None or str(raw_val).strip() == "":
                    is_error = True; break
                obj_val = str(raw_val)

                if col.startswith("_11_"):
                    if obj_val != true_elev: is_error = True; break
                elif col.startswith("_01_"):
                    if obj_val != true_space: is_error = True; break
                elif col.startswith("_05_"):
                    if obj_val != dim_w: is_error = True; break
                elif col.startswith("_06_"):
                    if obj_val != dim_d: is_error = True; break
                elif col.startswith("_07_"):
                    if obj_val != dim_h: is_error = True; break
                elif col.startswith("_CB."):
                    expected = str(tag_data.get(col, "-"))
                    if is_cb_layer:
                        if expected == "-":
                            if obj_val != "-": is_error = True; break
                        else:
                            if obj_val == "-": is_error = True; break
                    else:
                        if obj_val != "-": is_error = True; break
                else:
                    is_protected = any(col.startswith(wl[:4]) for wl in WHITE_LIST)
                    if not is_protected and obj_val != str(tag_data.get(col, "-")):
                        is_error = True; break

        if is_error: red_list.append(guid)

    rs.UnselectAllObjects()
    rs.EnableRedraw(True)

    if not red_list:
        rs.MessageBox(u"Global validation complete: all attributes match and all UUIDs are present!")
    else:
        rs.SelectObjects(red_list)
        rs.MessageBox(u"Issues found! Selected {} objects with attribute errors, missing/invalid/duplicate UUIDs.".format(len(red_list)))

# ==================================================================
# System Layer Management
# ==================================================================
def _ensure_system_layers():
    data_root = LAYER_PREFIX_3D + "::" + LAYER_DATA_SUFFIX
    if not rs.IsLayer(data_root):
        rs.AddLayer(data_root, COLOR_DATA_LAYER)
    rs.LayerColor(data_root, _to_color(COLOR_DATA_LAYER))
    for sys_layer in [LAYER_SPACE_BOUNDARIES, LAYER_LEVEL_FFL, LAYER_LEVEL_FL]:
        if not rs.IsLayer(sys_layer):
            rs.AddLayer(sys_layer, COLOR_DATA_LAYER)
        rs.LayerColor(sys_layer, _to_color(COLOR_DATA_LAYER))

# ==================================================================
# Rhino to XLSX: Reverse Sync
# ==================================================================
def func_rhino_to_xlsx():
    orig_path = get_dictionary_path(False)
    if not orig_path or not orig_path.lower().endswith('.xlsx'):
        rs.MessageBox(u"This function requires an existing .xlsx dictionary as a base and format reference.", 48)
        return

    try:
        wb = openpyxl.load_workbook(orig_path)
        ws = wb.active
        header_row = -1

        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if DICTIONARY_KEY_COLUMN in [str(c).strip() for c in row if c is not None]:
                header_row = i
                break

        if header_row == -1:
            rs.MessageBox(u"Reference dictionary is missing the '{}' header; cannot proceed with export.".format(DICTIONARY_KEY_COLUMN), 16)
            return

        col_map = {}
        for cell in ws[header_row]:
            if cell.value is not None:
                col_map[str(cell.value).strip()] = cell.column

        layer_col_idx = col_map.get(DICTIONARY_KEY_COLUMN)
        if not layer_col_idx: return

        raw_rhino_layers = [str(l) for l in rs.LayerNames() if str(l).startswith(_PREFIX_WITH_SEP)]

        export_full_layers = []
        for lname in raw_rhino_layers:
            if _is_data_layer(lname): continue
            if lname == LAYER_PREFIX_3D + "::" + LAYER_DATA_SUFFIX: continue
            if lname.startswith(LAYER_DW_PLAN + "::"): continue
            if lname == LAYER_DW_PLAN:
                export_full_layers.append(lname)
                continue
            is_parent = any(other.startswith(lname + "::") for other in raw_rhino_layers)
            if not is_parent:
                export_full_layers.append(lname)

        full_to_rel = {f: _strip_prefix(f) for f in export_full_layers}
        rel_to_full = {v: k for k, v in full_to_rel.items()}
        export_rel_set    = set(full_to_rel.values())
        processed_rel_set = set()

        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=layer_col_idx).value
            if not cell_val: continue

            raw_name   = str(cell_val).strip()
            clean_name = raw_name.replace(" [NEW]", "").replace(" [DELETED]", "").replace(" [MODIFIED]", "").replace(" [EXCLUDED]", "").strip()
            clean_rel  = _strip_prefix(clean_name)

            if clean_rel in export_rel_set:
                processed_rel_set.add(clean_rel)
                full_path = rel_to_full[clean_rel]
                l_idx = sc.doc.Layers.FindByFullPath(full_path, -1)
                if l_idx < 0: continue
                layer_obj = sc.doc.Layers[l_idx]

                is_modified = False
                for h_name, c_idx in col_map.items():
                    if h_name == DICTIONARY_KEY_COLUMN: continue
                    if h_name.startswith("_"):
                        rhino_val = layer_obj.GetUserString(h_name) or "-"
                        excel_val = str(ws.cell(row=row_idx, column=c_idx).value or "-").strip()
                        if rhino_val != excel_val:
                            ws.cell(row=row_idx, column=c_idx).value = rhino_val
                            is_modified = True

                ws.cell(row=row_idx, column=layer_col_idx).value = clean_rel + (" [MODIFIED]" if is_modified else "")
            else:
                excluded = clean_name in raw_rhino_layers or (_PREFIX_WITH_SEP + clean_rel) in raw_rhino_layers
                marker = " [EXCLUDED]" if excluded else " [DELETED]"
                ws.cell(row=row_idx, column=layer_col_idx).value = clean_rel + marker

        for rel_name in export_rel_set:
            if rel_name not in processed_rel_set:
                full_path = rel_to_full[rel_name]
                new_row   = ws.max_row + 1
                ws.cell(row=new_row, column=layer_col_idx).value = rel_name + " [NEW]"
                l_idx = sc.doc.Layers.FindByFullPath(full_path, -1)
                if l_idx >= 0:
                    layer_obj = sc.doc.Layers[l_idx]
                    for h_name, c_idx in col_map.items():
                        if h_name == DICTIONARY_KEY_COLUMN: continue
                        if h_name.startswith("_"):
                            ws.cell(row=new_row, column=c_idx).value = layer_obj.GetUserString(h_name) or "-"

        doc_dir   = rs.DocumentPath()
        if not doc_dir: doc_dir = os.path.expanduser("~")
        save_path = rs.SaveFileName(u"Save New Dictionary (preserve format)", "Excel (*.xlsx)|*.xlsx||", doc_dir, "LoopFlow_Dictionary_Export.xlsx")

        if save_path:
            if not save_path.lower().endswith(".xlsx"): save_path += ".xlsx"
            wb.save(save_path)
            rs.MessageBox(u"Reverse dictionary export complete!", 64)

    except Exception as e:
        if log_exception: print(log_exception(u"LF_Nexus.func_rhino_to_xlsx", e))
        rs.MessageBox(u"Exception during export:\n{}\n\nSee debug log: {}".format(e, _DEBUG_LOG_PATH), 16)

# ==================================================================
# UI Helpers
# ==================================================================
class _NexusChooserDialog(forms.Dialog[str]):
    """Eto-based replacement for rs.ListBox; width/height fully controllable."""
    def __init__(self, title, message, items, width=640, height=260):
        super().__init__()
        self.Title       = title
        self.ClientSize  = drawing.Size(width, height)
        self.Padding     = drawing.Padding(10)
        self.Resizable   = True
        self._result     = None
        self._items_text = list(items)

        prompt = forms.Label()
        prompt.Text = message

        self._listbox = forms.ListBox()
        self._listbox.Font = drawing.Font("Consolas", 10)
        for it in self._items_text:
            li = forms.ListItem()
            li.Text = it
            self._listbox.Items.Add(li)
        if self._items_text:
            self._listbox.SelectedIndex = 0
        self._listbox.MouseDoubleClick += self._on_ok

        ok_btn     = forms.Button()
        ok_btn.Text = "OK"
        cancel_btn = forms.Button()
        cancel_btn.Text = "Cancel"
        ok_btn.Click     += self._on_ok
        cancel_btn.Click += self._on_cancel

        layout = forms.DynamicLayout()
        layout.Spacing = drawing.Size(6, 6)
        layout.AddRow(prompt)
        layout.Add(self._listbox, yscale=True)

        btn_row = forms.DynamicLayout()
        btn_row.BeginHorizontal()
        btn_row.Add(None)
        btn_row.Add(ok_btn)
        btn_row.Add(cancel_btn)
        btn_row.EndHorizontal()
        layout.AddRow(btn_row)

        self.Content       = layout
        self.DefaultButton = ok_btn
        self.AbortButton   = cancel_btn

    def _on_ok(self, sender, e):
        idx = self._listbox.SelectedIndex
        if idx is not None and idx >= 0 and idx < len(self._items_text):
            self._result = self._items_text[idx]
        self.Close(self._result)

    def _on_cancel(self, sender, e):
        self._result = None
        self.Close(None)


def _show_nexus_chooser(items, title, message, width=640, height=260):
    """Mirror rs.ListBox signature: returns selected text, or None on cancel.
    Raises on Eto failure so that the caller can surface the real error
    instead of silently falling back to the narrow rs.ListBox.
    """
    dlg = _NexusChooserDialog(title, message, items, width, height)
    return dlg.ShowModal(Rhino.UI.RhinoEtoApp.MainWindow)


def Show_StatusInfo():
    info_text = u"New / Existing / Demolished / Relocated\n\n · _02_ Construction Status: overwrite this attribute value directly in the object's UserText to change individual status.\n\n · Default logic: [00_STR_Structure] layer defaults to Existing; all other layers default to New."
    rs.MessageBox(info_text, 0, u"Object Status Info")

# ==================================================================
# Dict to Layer: Build 3D Layer Structure
# ==================================================================
def func_dict_to_layer(df_dict):
    try:
        rs.EnableRedraw(False)

        color_map = {k: _to_color(v) for k, v in COLOR_LAYER_MAP.items()}
        _black = _to_color((0, 0, 0))

        recs = [r for r in df_dict.sort_values(by=DICTIONARY_KEY_COLUMN).to_dict(orient='records')
                if str(r[DICTIONARY_KEY_COLUMN]).strip() not in ["-", "nan", "None"]]

        if not recs: return

        _ensure_system_layers()

        for i, row in enumerate(recs):
            rel_path = str(row[DICTIONARY_KEY_COLUMN]).strip()
            fpath    = _build_full_layer_path(rel_path)

            if _is_data_layer(fpath):
                t_color = _to_color(COLOR_DATA_LAYER)
            else:
                t_color = color_map.get("furniture", _black) if "_Furniture" in fpath else _black
                for p in fpath.split("::"):
                    pre = p.split('_')[0]
                    if pre in color_map:
                        t_color = color_map[pre]
                        break

            path_parts  = fpath.split("::")
            current_path = ""
            for j, part in enumerate(path_parts):
                current_path = part if j == 0 else current_path + "::" + part
                if not rs.IsLayer(current_path):
                    rs.AddLayer(current_path, t_color)

            rs.LayerColor(fpath, t_color)
            l_idx = sc.doc.Layers.FindByFullPath(fpath, -1)
            if l_idx >= 0:
                layer_obj = sc.doc.Layers[l_idx]
                for col in df_dict.columns:
                    if col != DICTIONARY_KEY_COLUMN:
                        layer_obj.SetUserString(col, str(row.get(col, "-")))

                mat_idx = -1
                for m in sc.doc.Materials:
                    if m.Name == fpath and not m.IsDeleted:
                        mat_idx = m.Index; break
                if mat_idx == -1:
                    new_mat = Rhino.DocObjects.Material()
                    new_mat.Name = fpath
                    new_mat.DiffuseColor = t_color
                    new_mat.ToPhysicallyBased()
                    new_mat.PhysicallyBased.BaseColor = Rhino.Display.Color4f(t_color)
                    mat_idx = sc.doc.Materials.Add(new_mat)

                layer_obj.RenderMaterialIndex = mat_idx
                layer_obj.CommitChanges()

            y_coord = (len(recs) - 1 - i) * 10.0
            line_id = rs.AddLine([0, y_coord, 0], [-25, y_coord, 0])

            if line_id:
                rs.ObjectLayer(line_id, fpath)
                id_name_col = next((c for c in df_dict.columns if c.startswith("_04_")), None)
                ref_name = str(row.get(id_name_col, "-")) if id_name_col else "-"
                if ref_name == "-": ref_name = fpath.split("::")[-1]
                rs.ObjectName(line_id, "DNA_REF_" + ref_name)
                for col in df_dict.columns:
                    if col != DICTIONARY_KEY_COLUMN:
                        rs.SetUserText(line_id, col, str(row.get(col, "-")))

        rs.ZoomExtents(None, True)
        rs.MessageBox(u"Base layer structure built successfully!", 64)

    except Exception as e:
        if log_exception: print(log_exception(u"LF_Nexus.func_dict_to_layer", e))
        rs.MessageBox(u"Layer build error:\n{}\n\nSee debug log: {}".format(e, _DEBUG_LOG_PATH), 16)
    finally:
        rs.EnableRedraw(True)

# ==================================================================
# Space Boundary Setter
# ==================================================================
def func_boundary_setter():
    if not rs.IsLayer(LAYER_SPACE_BOUNDARIES):
        rs.AddLayer(LAYER_SPACE_BOUNDARIES, (0, 255, 255))
    crv_id = rs.GetObject(u"Select space boundary (closed curve)", rs.filter.curve)
    if crv_id and rs.IsCurveClosed(crv_id):
        old_name = rs.GetUserText(crv_id, "Space_Name") or ""
        new_name = rs.StringBox(u"Enter space name:", old_name, u"Define Space")
        if new_name:
            rs.SetUserText(crv_id, "Space_Name", new_name)
            rs.ObjectName(crv_id, "Zone_[{}]".format(new_name))
            rs.ObjectLayer(crv_id, LAYER_SPACE_BOUNDARIES)

def main():
    try:
        options = [
            u"▷ Dict. to Layer  | Load LoopFlow dictionary and build layers",
            u"▶ SpaceBoundary   | Define Space Boundary",
            u"▶ TagTrigger       | Write attributes globally (includes hidden/locked)",
            u"▶ TagChecker       | Validate attributes globally for errors and missing values",
            u"▷ Layer to Dict.  | Reverse-sync changes back to XLSX",
            u"|| Object Status Info ||"
        ]
        # Dialog size is controlled directly here (width, height in pixels).
        choice = _show_nexus_chooser(options, u"Nexus", u"Select function:", width=620, height=220)
        if not choice: return

        if u"Dict. to Layer" in choice:
            dict_path = get_dictionary_path(force_select=True)
            if not dict_path: return
            df_dict = load_dict_from_path(dict_path)
            if df_dict is not None:
                func_dict_to_layer(df_dict)

        elif u"TagTrigger" in choice or u"TagChecker" in choice:
            dict_path = get_dictionary_path(force_select=False)
            if not dict_path: return
            df_dict = load_dict_from_path(dict_path)
            if df_dict is None: return

            if u"TagTrigger" in choice:
                func_tag_trigger(df_dict)
            else:
                func_tag_checker(df_dict)

        elif u"SpaceBoundary" in choice:
            func_boundary_setter()
        elif u"Layer to Dict." in choice:
            func_rhino_to_xlsx()
        elif u"Object Status Info" in choice:
            Show_StatusInfo()

    except Exception as e:
        if log_exception:
            print(log_exception(u"LF_Nexus.main", e))
        rs.MessageBox(u"LF_Nexus encountered an unexpected error.\n\nSee debug log: {}".format(_DEBUG_LOG_PATH), 16)

if __name__ == "__main__":
    main()
